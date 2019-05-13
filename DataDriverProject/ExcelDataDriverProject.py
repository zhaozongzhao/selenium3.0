from openpyxl import load_workbook
class ParseExcel(object):

    def __init__(self,excelpath,sheetName):
        # 将要读取的Excel加载到内存中
        self.wd = load_workbook(excelpath)
        #获取指定sheet的对象
        self.sheet = self.wd[sheetName]
        print(self.sheet.title)
        #获取最大行数
        self.maxRowNum = self.sheet.max_row

    def getDataFromeSheet(self):
        #用于存放工作 表中读取的数据
        datalist = []
        # 循环获取当前所有的行数对象
        for line in self.sheet.rows:
            # 获取每行中每个的数值
            tmplist = []
            tmplist.append(line[1].value)
            tmplist.append(line[2].value)
            datalist.append(tmplist)
        return datalist

    def get_caseDatas_all(self):
        # 用于存放工作 表中读取的数据
        all_case_datas = []
        # 循环获取最大行的行数，从第2行开始
        for index in range(2, self.sheet.max_row + 1):
            case_data = []
            # sheet.cell 根据输入等列和行，获取单元格的数值
            case_data.append(self.sheet.cell(index, 2).value)
            case_data.append(self.sheet.cell(index, 3).value)
            all_case_datas.append(case_data)
        return all_case_datas
if __name__=='__main__':

    for i in pe.getDataFromeSheet():
        print(i[0],i[1])
